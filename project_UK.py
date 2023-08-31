import numpy_financial as npf
import numpy as np
import pandas as pd
from openpyxl.styles import Border, Side, Font, PatternFill
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

print('\nPURCHASE DATA\n')
purchase_p = float(input('Purchase price is [€]: '))
while purchase_p == 0:
    print('Value must be greater than 0. Please enter a valid value')
    purchase_p = float(input('Purchase price is [€]: '))

down_p = float(input('Down payment is [%]: ')) / 100
while down_p == 0:
    print('Value must be greater than 0. Please enter a valid value')
    down_p = float(input('Down payment is [%]: ')) / 100

i_rate = float(input('Interest rate is [%]: '))/100
while i_rate == 0:
    print('Value must be greater than 0. Please enter a valid value')
    i_rate = float(input('Interest rate is [%]: '))/100

loan_t = int(input('Loan term is [years]: '))
while loan_t == 0:
    print('Value must be greater than 0. Please enter a valid value')
    loan_t = int(input('Loan term is [years]: '))

closing_c = float(input('Closing cost is [€]: '))
repair_c = float(input('Repair Cost is [€]: '))


def price_ar():
    if repair_c == 0:
        price_after_repair = 0
    else:
        price_after_repair = purchase_p + purchase_p*down_p + repair_c
    return price_after_repair


print('\nRECURRING OPERATING EXPENSES PER YEAR\n')
property_tax = float(input('Property tax is [€/year]: '))
property_tax_pro = float(input('Annual increase of property tax is [%]: '))/100
if property_tax_pro == 0:
    property_tax_pro = float(1/100)

insurance = float(input('Total insurance is [€/year]: '))
insurance_pro = float(input('Annual increase of insurance is [%]: '))/100
if insurance_pro == 0:
    insurance_pro = float(1/100)

HOA_fee = float(input('HOA fee is [€/year]: '))
HOA_fee_pro = float(input('Annual increase of HOA fee is [%]: '))/100
if HOA_fee_pro == 0:
    HOA_fee_pro = float(1/100)

maintenance = float(input('Maintenance is [€/year]: '))
maintenance_pro = float(input('Annual increase of maintenance is [%]: '))/100
if maintenance_pro == 0:
    maintenance_pro = float(1/100)

other_costs = float(input('Other expenses is [€/year]: '))
other_costs_pro = float(input('Annual increase of other expenses is [%]: '))/100
if other_costs_pro == 0:
    other_costs_pro = float(1/100)

print('\nINCOME PER MONTH\n')
rent = float(input('Monthly rent income is [€]: '))
rent_pro = float(input('Annual increase of rent income is [%]: '))/100
if rent_pro == 0:
    rent_pro = float(1/100)

other_income = float(input('Other monthly income is [€]: '))
other_income_pro = float(input('Annual increase of other income is [%]: '))/100
if other_income_pro == 0:
    other_income_pro = float(1/100)

vacancy_rate = float(input('Vacancy rate is [%]: '))/100
management_fee = float(input('Management fee is [%]: '))/100

print('\nSALE\n')
v_appreciation = float(input('Yearly value appreciation is [%]: '))/100
holding_length = int(input('Holding length is [years]: '))
cost_sell = float(input('Cost to sell is [%]: '))/100


# #################TABLE 1 (Revenue and expenditure in the first year)#################
# INCOME
def income():
    return float(rent) + float(other_income)


def income_y():
    return income()*12


# MORTGAGE PAYMENT
def mortgage_mpt():
    return -npf.pmt(i_rate/12, loan_t*12, purchase_p-purchase_p*down_p, 0)


def mortgage_mpt_y():
    return mortgage_mpt()*12


# VACANCY
def vacancy():
    return rent * vacancy_rate + other_income * vacancy_rate


def vacancy_y():
    return vacancy()*12


# MANAGEMENT FEE
def management():
    return (rent - rent * vacancy_rate) * management_fee\
        + (other_income - other_income * vacancy_rate) * management_fee


def management_y():
    return management()*12


# PROPERTY TAX
def pro_tax():
    return property_tax/12


def pro_tax_y():
    return pro_tax()*12


# TOTAL INSURANCE
def ins():
    return insurance/12


def ins_y():
    return ins()*12


# HOA FEE
def hoa():
    return HOA_fee/12


def hoa_y():
    return hoa()*12


# MAINTENANCE FEE
def maint():
    return maintenance/12


def maint_y():
    return maint()*12


# OTHER COSTS
def other_c():
    return other_costs/12


def other_c_y():
    return other_c()*12


# CASH FLOW
def cf_1y():
    return income()-mortgage_mpt()-vacancy()-management()-pro_tax()-ins()-hoa()-maint()-other_c()


# print(cf_1y())
def cf_1y_y():
    return cf_1y()*12


# NET OPERATION INCOME (NOI):
def noi():
    return income() - vacancy() - management() - pro_tax() - ins() - hoa() - maint() - other_c()


def noi_y():
    return noi()*12


# Table 1
table_1 = {'': ['Income:',
                'Mortgage Pay:',
                f'Vacancy {int(vacancy_rate*100)} %:',
                f'Management Fee {int(management_fee*100)} %:',
                'Property Tax:',
                'Total Insurance:', 'HOA Fee:',
                'Maintenance Cost:', 'Other Cost:',
                'Cash Flow:',
                'Net Operating Income (NOI):'],
           'Monthly': [f'{round(income(),2)} €',
                       f'{round(mortgage_mpt(),2)} €',
                       f'{round(vacancy(),2)} €',
                       f'{round(management(), 2)} €',
                       f'{round(pro_tax(), 2)} €',
                       f'{round(ins(), 2)} €',
                       f'{round(hoa(), 2)} €',
                       f'{round(maint(), 2)} €',
                       f'{round(other_c(), 2)} €',
                       f'{round(cf_1y(), 2)} €',
                       f'{round(noi(), 2)} €'],
           'Annual': [f'{round(income_y(), 2)} €',
                      f'{round(mortgage_mpt_y(), 2)} €',
                      f'{round(vacancy_y(), 2)} €',
                      f'{round(management_y(), 2)} €',
                      f'{round(pro_tax_y(), 2)} €',
                      f'{round(ins_y(), 2)} €',
                      f'{round(hoa_y(), 2)} €',
                      f'{round(maint_y(), 2)} €',
                      f'{round(other_c_y(), 2)} €',
                      f'{round(cf_1y_y(), 2)} €',
                      f'{round(noi_y(), 2)} €']}

df_1 = pd.DataFrame(table_1)
styled_table_1 = df_1.style

# Text alignment
styled_table_1.set_properties(**{'text-align': 'center'})


# #################TABLE 2 (Breakdown Over Time)#################
# INCOME
def income_repeat():
    i = 0
    list_1 = []
    while i < holding_length:
        i += 1
        v1 = round((rent*12*(1+rent_pro)**(i-1)
                    - rent*(1+rent_pro)**(i-1)*vacancy_rate*12
                    - (rent*12*(1+rent_pro)**(i-1)-rent*(1+rent_pro)**(i-1)*vacancy_rate*12)*management_fee
                    + other_income*12*(1+other_income_pro)**(i-1)
                    - other_income*(1+other_income_pro)**(i - 1)*vacancy_rate*12
                    - (other_income*12*(1+other_income_pro)**(i-1)
                    - other_income*(1+other_income_pro)**(i - 1)*vacancy_rate*12)*management_fee))

        list_1.append(v1)
    return list_1


# MORTGAGE
def mortgage_repeat():
    list_2 = []
    i = 0
    while i < holding_length:
        i += 1
        v2 = round(mortgage_mpt_y())
        list_2.append(v2)
    return list_2


# EXPENSES
def expenses_repeat():
    list_3 = []
    i = 0
    while i < holding_length:
        i += 1
        v3 = round((pro_tax_y()*(1+property_tax_pro)**(i-1))
                   + ins_y()*((1+insurance_pro)**(i-1))
                   + hoa_y()*((1+HOA_fee_pro)**(i-1))
                   + maint_y()*((1+maintenance_pro)**(i-1))
                   + other_c_y()*((1+other_costs_pro)**(i-1)))
        list_3.append(v3)
    return list_3


# EXPENSES IN 1 YEAR
def fy_expenses():
    return -round(purchase_p*down_p+closing_c+repair_c)


# CASH FLOW FOR EACH YEAR
def cf_repeat():
    list_1 = income_repeat()
    list_2 = mortgage_repeat()
    list_3 = expenses_repeat()
    list_4 = [x - y - z for x, y, z in zip(list_1, list_2, list_3)]
    return list_4


# CASH ON CASH RETURN FOR EACH YEAR
def coc_repeat():
    list_4 = cf_repeat()
    d = fy_expenses()
    list_5 = [round(x*100/-d, 2) for x in list_4]
    return list_5


# EQUITY ACCUMULATED
def cumipmt(rate, nper, pv, start_period, end_period, payment_type):
    pmt = -npf.pmt(rate, nper, pv, 0, payment_type)
    interest_paid = 0.0
    for period in range(start_period, end_period + 1):
        ipmt = pv * rate
        ppmt = pmt - ipmt
        pv -= ppmt
        interest_paid += ipmt
    return interest_paid


def ac_repeat():
    list_6 = []
    for i in range(holding_length):
        if repair_c == 0:
            v6 = round(purchase_p * down_p
                       + (mortgage_mpt_y() * (i + 1)
                          - cumipmt(i_rate/12, loan_t * 12, purchase_p * (1 - down_p), 1, 12 * (i + 1), 0))
                       + (purchase_p * (1 + v_appreciation) ** (i + 1) - purchase_p))
        else:
            v6 = round(purchase_p * down_p
                       + (mortgage_mpt_y() * (i + 1)
                          - cumipmt(i_rate/12, loan_t * 12, purchase_p * (1 - down_p), 1, 12 * (i + 1), 0))
                       + (price_ar() * (1 + v_appreciation) ** (i+1) - purchase_p))

        list_6.append(v6)
    return list_6


# CASH TO RECEIVE:
def ctr():
    list_6 = ac_repeat()
    list_7 = []

    for i in range(holding_length):
        if repair_c == 0:
            subtract = purchase_p * (1 + v_appreciation) ** (i + 1) * cost_sell
        else:
            subtract = price_ar() * (1 + v_appreciation) ** (i + 1) * cost_sell

        result = round(list_6[i] - subtract)
        list_7.append(result)

    return list_7


# IRR
cash_flows = []
expenses = []
cash_to_receive = []
cash_flows.extend(cf_repeat())
expenses.append(fy_expenses())
cash_to_receive.extend(ctr())


def calculate_irr():
    irrs = []
    for y in range(1, holding_length+1):
        cash_flow = [expenses[0]] + cash_flows[:y-1] + [cash_to_receive[y-1] + cash_flows[y-1]]
        irr = round(npf.irr(cash_flow)*100, 2)
        irrs.append(irr)
    return irrs


# TABLE_2
def each_holding_y():
    years = []
    for y in range(1, holding_length + 1):
        years.append(y)
    return years


# Table 2
df_initial = pd.DataFrame({'Years': each_holding_y(),
                           'Annual Income [€]': income_repeat(),
                           'Mortgage [€]': mortgage_repeat(),
                           'Expenses [€]': expenses_repeat(),
                           'Cash Flow [€]': cf_repeat(),
                           'Cash on Cash Return [%]': coc_repeat(),
                           'Equity Accumulated [€]': ac_repeat(),
                           'Cash to Receive [€]': ctr(),
                           'Return [%]': calculate_irr()})

new_row = pd.DataFrame({'Cash Flow [€]': fy_expenses()}, index=[0])
df_concatenated = pd.concat([new_row, df_initial], ignore_index=True)
df_concatenated.insert(4, 'Cash Flow [€]', df_concatenated.pop('Cash Flow [€]'))

totals = df_concatenated[['Annual Income [€]', 'Mortgage [€]', 'Expenses [€]', 'Cash Flow [€]']].sum()
totals_row = pd.DataFrame({'Years': 'Total',
                           'Annual Income [€]': totals['Annual Income [€]'],
                           'Mortgage [€]': totals['Mortgage [€]'],
                           'Expenses [€]': totals['Expenses [€]'],
                           'Cash Flow [€]':
                               totals['Cash Flow [€]']
                               + df_concatenated.loc[df_concatenated['Years']
                                                     == each_holding_y()[-1], 'Cash to Receive [€]'].values[0],
                           'Cash on Cash Return [%]': round((totals['Cash Flow [€]']
                                                             + df_concatenated.loc[df_concatenated['Years']
                                                             == each_holding_y()[-1], 'Cash to Receive [€]']
                                                             .values[0])/-fy_expenses()*100, 2)}, index=[len(df_1)])

df_2 = pd.concat([df_concatenated, totals_row]).reset_index(drop=True)

df_2 = pd.DataFrame(df_2)
styled_table_2 = df_2.style

# Text alignment
styled_table_2.set_properties(**{'text-align': 'center'})


# #################TABLE 3 (f'For {holding_length} years invested')#################
# RETURN (IRR)
def irr_last_year():
    cash_flow_last_year = [expenses[0]] + cash_flows[:holding_length-1] + [cash_to_receive[-1]+cash_flows[-1]]
    irr_ly = npf.irr(cash_flow_last_year) * 100
    return irr_ly


# TOTAL PROFIT WHEN SOLD
def total_profit_when_sold():
    tpws = [float(expenses[0])] + cash_flows[:holding_length] + [float(cash_to_receive[-1])]
    return sum(tpws)


# CASH ON CASH RETURN
def coc_return_total():
    coc_rt = total_profit_when_sold()/-expenses[0]*100
    return coc_rt


# CAPITALIZATION RATE
def cap_rate():
    return round(noi_y()/purchase_p * 100, 2)


# TOTAL RENTAL INCOME
def total_rental_income():
    fv_rent = npf.fv(rent_pro, holding_length,
                     - (rent * 12 - rent * 12 * vacancy_rate
                        - (rent * 12 - rent * 12 * vacancy_rate) * management_fee), 0)
    fv_other_income = npf.fv(other_income_pro, holding_length,
                             - (other_income * 12 - other_income * 12 * vacancy_rate
                                - (other_income * 12 - other_income * 12 * vacancy_rate) * management_fee), 0)
    return fv_rent + fv_other_income


# TOTAL MORTGAGE PAYMENT
def total_mortgage_payment():
    if loan_t < holding_length:
        return round(mortgage_mpt_y()*loan_t, 2)
    else:
        return round(mortgage_mpt_y()*holding_length, 2)


# TOTAL EXPENSES
def total_expenses():
    fv_pt = npf.fv(property_tax_pro, holding_length, -property_tax, 0)
    fv_ti = npf.fv(insurance_pro, holding_length, -insurance, 0)
    fv_hoa = npf.fv(HOA_fee_pro, holding_length, -HOA_fee, 0)
    fv_maint = npf.fv(maintenance_pro, holding_length, -maintenance, 0)
    fv_other_costs = npf.fv(other_costs_pro, holding_length, -other_costs, 0)
    return round(fv_pt + fv_ti + fv_hoa + fv_maint + fv_other_costs, 2)


# TOTAL NET OPERATION INCOME
def total_noi():
    return round(total_rental_income()-total_expenses(), 2)


# table 3
table_3 = {'': ['IRR:',
                'Total Profit:',
                'Cash on Cash Return:',
                'Capitalization Rate:',
                'Total Rental Income:',
                'Total Mortgage Payments:',
                'Total Expenses:',
                'Total Net Operating Income:'],
           'Result': [f'{round(irr_last_year(), 2)} %',
                      f'{round(total_profit_when_sold(), 2)} €',
                      f'{round(coc_return_total(), 2)} %',
                      f'{round(cap_rate(), 2)} %',
                      f'{round(total_rental_income(), 2)} €',
                      f'{round(total_mortgage_payment(), 2)} €',
                      f'{round(total_expenses(), 2)} €',
                      f'{round(total_noi(), 2)} €']}


df_3 = pd.DataFrame(table_3)
styled_table_3 = df_3.style

# Text alignment
styled_table_3.set_properties(**{'text-align': 'center'})

# #################FINAL TABLE TO SAVE #################
writer = pd.ExcelWriter('Calculation_result.xlsx', engine='openpyxl')
workbook = writer.book

styled_table_1.to_excel(writer, sheet_name='Sheet1', index=False, startcol=1, startrow=3)
styled_table_2.to_excel(writer, sheet_name='Sheet2', index=False, startcol=1, startrow=3)
styled_table_3.to_excel(writer, sheet_name='Sheet3', index=False, startcol=1, startrow=3)

sheet1 = workbook['Sheet1']
sheet2 = workbook['Sheet2']
sheet3 = workbook['Sheet3']

# Cells coloring
dark_fill = PatternFill(start_color='c6bcb6', end_color='c6bcb6', fill_type='solid')
light_fill = PatternFill(start_color='e4d1d1', end_color='e4d1d1', fill_type='solid')

for row in sheet1.iter_rows(min_row=5, min_col=2):
    for cell in row:
        cell.fill = light_fill

for cell in sheet1.iter_cols(min_row=4, min_col=2):
    cell[0].fill = dark_fill

for row in sheet2.iter_rows(min_row=5, min_col=2):
    for cell in row:
        cell.fill = light_fill

for cell in sheet2.iter_cols(min_row=4, min_col=2):
    cell[0].fill = dark_fill

for row in sheet3.iter_rows(min_row=5, min_col=2):
    for cell in row:
        cell.fill = light_fill

for cell in sheet3.iter_cols(min_row=4, min_col=2):
    cell[0].fill = dark_fill

# Border line
border_style = Border(left=Side(style='thin'),
                      right=Side(style='thin'),
                      top=Side(style='thin'),
                      bottom=Side(style='thin'))

for row in sheet1.iter_rows(min_row=4, min_col=2):
    for cell in row:
        cell.border = border_style

for row in sheet2.iter_rows(min_row=4, min_col=2):
    for cell in row:
        cell.border = border_style

for row in sheet3.iter_rows(min_row=4, min_col=2):
    for cell in row:
        cell.border = border_style

# Cell width
for column_cells in sheet1.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        if cell.value:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
    adjusted_width = (max_length + 2) * 1.2
    sheet1.column_dimensions[column].width = adjusted_width

for column_cells in sheet2.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        if cell.value:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
    adjusted_width = (max_length + 2) * 1.2
    sheet2.column_dimensions[column].width = adjusted_width

for column_cells in sheet3.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        if cell.value:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
    adjusted_width = (max_length + 2) * 1.2
    sheet3.column_dimensions[column].width = adjusted_width

# Bold font
bold_font = Font(bold=True)

for column_cells in sheet1.iter_cols(min_row=14, max_row=14):
    for cell in column_cells:
        cell.font = bold_font

total_row_index = df_2[df_2['Years'] == 'Total'].index[0]+3
for row_num, row in enumerate(sheet2.iter_rows(min_row=total_row_index+2, max_row=total_row_index+2)):
    for cell in row:
        cell.font = bold_font

for column_cells in sheet3.iter_cols(min_row=5, max_row=8):
    for cell in column_cells:
        cell.font = bold_font

bold_font = Font(bold=True, italic=True, size=14)

sheet1.merge_cells('B2:D2')
sheet1['B2'].value = 'Revenue and expenditure in the first year'
sheet1['B2'].font = bold_font

sheet2.merge_cells('B2:D2')
sheet2['B2'].value = 'Breakdown Over Time'
sheet2['B2'].font = bold_font

sheet3.merge_cells('B2:D2')
sheet3['B2'].value = f'For {holding_length} years invested'
sheet3['B2'].font = bold_font

writer.save()

# PLOTS
# Plot for IRR
years_plot = df_2['Years'].tolist()
irr_values = df_2['Return [%]'].tolist()

valid_indices = np.where(~np.isnan(irr_values))[0]
years_plot = [years_plot[i] for i in valid_indices]
irr_values = [irr_values[i] for i in valid_indices]

plt.plot(years_plot, irr_values, marker='o')
plt.xlabel('Years')
plt.ylabel('IRR (%)')
plt.title('Changes in IRR values over the years ', fontweight='bold', fontsize=12)
plt.grid(True)

plt.xticks(years_plot[::1])

plt.savefig('plot_2.png')

total_rows = len(df_2)
image_row = total_rows + 6

workbook = load_workbook('Calculation_result.xlsx')
worksheet = workbook['Sheet2']

img = Image('plot_2.png')
worksheet.add_image(img, f'A{image_row}')

workbook.save('Calculation_result.xlsx')

# Plot for expenses
values = df_1['Annual'][1:-2].str.extract(r'(\d+\.\d+)').astype(float).values.flatten()
labels = [label[:-1] for label in df_1[''][1:-2].tolist()]

nonzero_values = [v for v, l in zip(values, labels) if v != 0]
nonzero_labels = [l for v, l in zip(values, labels) if v != 0]

fig, ax = plt.subplots(figsize=(5, 3), subplot_kw=dict(aspect="equal"))
ax.pie(nonzero_values, labels=nonzero_labels, autopct='%1.0f%%', textprops=dict(color="w", weight="bold"))

legend = ax.legend(bbox_to_anchor=(1, 0, 0.5, 0.5), prop={'size': 6})
plt.title('Breakdown of expenses in the first year', fontweight='bold', fontsize=12)

plt.savefig('plot.png')

workbook = load_workbook('Calculation_result.xlsx')
worksheet = workbook['Sheet1']

img = Image('plot.png')
worksheet.add_image(img, 'B17')

workbook.save('Calculation_result.xlsx')
